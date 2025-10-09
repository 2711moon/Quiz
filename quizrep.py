import time
import openpyxl 
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException,
    ElementClickInterceptedException, NoAlertPresentException
)
from datetime import datetime
from openpyxl import Workbook

# -------------- CONFIG -------------- #
QUIZ_URL = "https://training.kisna.com/"
DEPARTMENT_TO_SELECT = "IT"
MAX_WAIT = 15
EXCEL_FILE = "users.xlsx"  #excel file with Names and respective Mobiles
# ------------------------------------ #


def safe_click(driver, element, retries=2):
    """Try to click an element safely with retries."""
    for attempt in range(retries):
        try:
            element.click()
            return True
        except ElementClickInterceptedException:
            driver.execute_script("arguments[0].scrollIntoView();", element)
            time.sleep(0.5)
        except Exception:
            time.sleep(0.5)
    return False


def load_users_from_excel(filename):
    """Read Name and Mobile from users.xlsx"""
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    users = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header row
        name, mobile = row
        if name and mobile:
            users.append((str(name), str(mobile)))
    return users


def save_report_to_excel(attempted, already_done, skipped):
    """Save final report to an Excel file with 4 sheets and timestamp."""
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"quiz_report_{timestamp}.xlsx"

    wb = Workbook()

    #Attempted
    ws1 = wb.active
    ws1.title = "Attempted"
    ws1.append(["Name"])
    for name in attempted:
        ws1.append([name])

    #Already done
    ws2 = wb.create_sheet("Already Done")
    ws2.append(["Name"])
    for name in already_done:
        ws2.append([name])

    #Skipped
    ws3 = wb.create_sheet("Skipped")
    ws3.append(["Name"])
    for name in skipped:
        ws3.append([name])

    #Summary
    ws4 = wb.create_sheet("Summary")
    ws4.append(["Category", "Count"])
    ws4.append(["Attempted", len(attempted)])
    ws4.append(["Already Done", len(already_done)])
    ws4.append(["Skipped", len(skipped)])
    ws4.append(["Total", len(attempted) + len(already_done) + len(skipped)])

    wb.save(filename)
    print(f"\n📊 Excel report saved as '{filename}'\n")


def run_quiz_for_user(driver, wait, name, mobile, chosen_index, ans_word):
    print(f"\n🚀 Running quiz for {name} ({mobile})")

    try:
        back_btn = wait.until(EC.element_to_be_clickable((By.XPATH,
            "//a[contains(., 'BACK OFFICE LOGIN') or contains(., 'Back Office')]")))
        safe_click(driver, back_btn)
        print("Clicked BACK OFFICE LOGIN")
    except Exception as e:
        print("❌ Could not click BACK OFFICE LOGIN:", e)
        return "skipped"

    time.sleep(1)

    #Department
    try:
        dept_select_el = wait.until(EC.presence_of_element_located((By.TAG_NAME, "select")))
        select = Select(dept_select_el)
        try:
            select.select_by_visible_text(DEPARTMENT_TO_SELECT)
        except Exception:
            select.select_by_index(0)
        print("Department selected:", DEPARTMENT_TO_SELECT)
    except Exception as e:
        print("❌ Department select failed:", e)

    #Credentials
    def find_input(xpaths):
        for xp in xpaths:
            try:
                return driver.find_element(By.XPATH, xp)
            except NoSuchElementException:
                continue
        return None

    name_input = find_input([
        "//input[@name='name']", "//input[@id='name']",
        "//input[contains(@placeholder, 'Name')]"
    ])
    phone_input = find_input([
        "//input[@name='mobile']", "//input[@id='mobile']",
        "//input[contains(@placeholder, 'Mobile')]"
    ])

    if name_input:
        name_input.clear()
        name_input.send_keys(name)
    if phone_input:
        phone_input.clear()
        phone_input.send_keys(mobile)
    print("Filled credentials.")

    try:
        cont = wait.until(EC.element_to_be_clickable((By.XPATH,
            "//button[contains(., 'Continue') or contains(., 'Proceed')]")))
        safe_click(driver, cont)
    except Exception:
        print("Continue button not found.")

    time.sleep(2)

    #Quiz selection
    try:
        selects = driver.find_elements(By.TAG_NAME, "select")
        quiz_select = None
        for sel in selects:
            s = Select(sel)
            if any("select" not in opt.text.lower() for opt in s.options):
                quiz_select = s
                break

        if not quiz_select:
            print("❌ No quiz dropdown found.")
            return "skipped"

        for opt in quiz_select.options:
            if "select" not in opt.text.lower():
                quiz_select.select_by_visible_text(opt.text)
                print("Selected quiz:", opt.text.strip())
                break

        time.sleep(1)

        #Alert if the quiz was already attempted
        try:
            alert = driver.switch_to.alert
            alert_text = alert.text
            print(f"⚠️ Popup for {name}: {alert_text}")
            alert.accept()
            try:
                logout_btn = driver.find_element(By.XPATH, "//a[contains(., 'Logout')] | //button[contains(., 'Logout')]")
                safe_click(driver, logout_btn)
                print(f"⏭️ Skipped {name} due to popup (already attempted).")
            except Exception:
                print("Logout not found after popup.")
            return "already_done"
        except NoAlertPresentException:
            pass

    except Exception:
        print("Quiz selection failed.")
        return "skipped"

    #Start the Quiz
    try:
        start_btn = wait.until(EC.element_to_be_clickable((By.XPATH,
            "//*[contains(text(),'Start Quiz') or contains(text(),'Start')]")))
        safe_click(driver, start_btn)
        print("Started quiz.")
    except Exception:
        print("Start Quiz button not found.")
    time.sleep(2)

    #Questions' loop
    qn = 0
    try:
        while True:
            option_elements = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".option_list .option"))
            )
            if not option_elements:
                break

            if chosen_index < len(option_elements):
                chosen = option_elements[chosen_index]
            else:
                chosen = option_elements[-1]
                print(f"⚠️ Only {len(option_elements)} options found. Falling back to last option.")

            safe_click(driver, chosen)
            print(f"Q{qn+1}: clicked {ans_word} option")

            try:
                nav_btns = WebDriverWait(driver, 5).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".next_btn, .skip_btn"))
                )
                next_btn = None
                skip_btn = None
                for b in nav_btns:
                    if "next" in b.get_attribute("class"):
                        next_btn = b
                    elif "skip" in b.get_attribute("class"):
                        skip_btn = b

                btn_to_click = next_btn or skip_btn
                if not btn_to_click or not safe_click(driver, btn_to_click, retries=3):
                    print("⚠️ Failed to click Next/Skip, breaking.")
                    break

                print("➡️ Moved to next question.")
            except TimeoutException:
                print("⚠️ No Next/Skip found. Probably last question.")
                break

            qn += 1
            time.sleep(1)

    except TimeoutException:
        print("❌ Timed out waiting for question.")
        return "skipped"

    # Logout
    try:
        logout_btn = driver.find_element(By.XPATH, "//a[contains(., 'Logout')] | //button[contains(., 'Logout')]")
        safe_click(driver, logout_btn)
        print("Logged out.")
    except Exception:
        print("Logout not found.")

    print(f"✅ Quiz finished for {name}. Questions answered: {qn}")
    time.sleep(2)
    return "attempted"


def main():
    ans_word = input("Type your answer for all questions (first/second/third/fourth): ").strip().lower()
    mapping = {"first": 1, "second": 2, "third": 3, "fourth": 4}
    chosen_index = mapping.get(ans_word, 1) - 1
    print(f"✅ Will use '{ans_word}' option for all questions.")

    users = load_users_from_excel(EXCEL_FILE)
    print(f"Found {len(users)} users in Excel.")

    attempted, already_done, skipped = [], [], []

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, MAX_WAIT)

    try:
        for name, mobile in users:
            driver.get(QUIZ_URL)
            result = run_quiz_for_user(driver, wait, name, mobile, chosen_index, ans_word)

            if result == "attempted":
                attempted.append(name)
            elif result == "already_done":
                already_done.append(name)
            else:
                skipped.append(name)

    finally:
        print("\n\n============== QUIZ REPORT ==============")
        print(f"✅ Attempted by automation: {len(attempted)} users")
        if attempted:
            print("   →", ", ".join(attempted))

        print(f"⚠️ Already attempted (skipped): {len(already_done)} users")
        if already_done:
            print("   →", ", ".join(already_done))

        print(f"🚫 Skipped before completion: {len(skipped)} users")
        if skipped:
            print("   →", ", ".join(skipped))

        print("----------------------------------------")
        print(f"Total users: {len(users)}")
        print("========================================\n")

        #Excel summary (Save the file)
        save_report_to_excel(attempted, already_done, skipped)

        driver.quit()


if __name__ == "__main__":
    main()
